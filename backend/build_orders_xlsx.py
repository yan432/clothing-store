"""Combine the 4 standardization CSVs into a single .xlsx with 4 sheets."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXPORTS = Path.home() / "Desktop/clothing-store/exports"

SHEETS = [
    ("Clean Customers",        EXPORTS / "clean_customers.csv",        "DCE6F1"),
    ("B2B Showroom",           EXPORTS / "b2b_orders.csv",             "F4D7D7"),
    ("Multi-customer Review",  EXPORTS / "multi_customer_review.csv",  "FFE9B0"),
    ("Orphans",                EXPORTS / "orphans.csv",                "E4E4E4"),
]

wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, size=11)
header_align = Alignment(horizontal="left", vertical="center")

for sheet_name, csv_path, color in SHEETS:
    ws = wb.create_sheet(title=sheet_name)
    if not csv_path.exists():
        continue
    with open(csv_path, "r", encoding="utf-8") as fp:
        reader = csv.reader(fp)
        for i, row in enumerate(reader):
            ws.append(row)
            if i == 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.alignment = header_align
    # auto-width (rough)
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = str(cell.value or "")
                if len(v) > max_len:
                    max_len = len(v)
            except: pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)
    ws.freeze_panes = "A2"

out_path = EXPORTS / "orders_standardized.xlsx"
wb.save(out_path)
print(f"✓ {out_path}")
